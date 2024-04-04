import openpyxl
import random
import string

def generate_code():
    # Generate a random 8-digit alphanumeric code
    code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    return code

# Load the Excel file
wb = openpyxl.load_workbook('Total Students submission.xlsx')
ws = wb.active

# Add a new column header for the unique code after the 7th column
new_column_index = 8
ws.cell(row=1, column=new_column_index, value="code")

# Iterate through each row in the Excel file
for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    # Generate a unique code
    unique_code = generate_code()
    # Store the unique code in the new column
    ws.cell(row=row_number, column=new_column_index, value=unique_code)

# Save the updated Excel file
wb.save('Total Students submission.xlsx')
